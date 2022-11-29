# PART 1: WORKING WITH DATA

import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook

# Importing original table:
df = pd.read_excel('dataset.xls')
df = df.drop('good (1)', axis=1)

# Removing duplicates in each area:
df = df.drop_duplicates(subset=['area', 'keyword'])

# Deleting empty rows (the ones without clusters):
df['cluster'] = pd.to_numeric(df['cluster'], errors='coerce')
df = df.dropna(subset=['cluster'])

# Adding 'color' column
df['color'] = ''

# Sorting by 'area', 'cluster', 'cluster_name' and 'count'(descending):
df['count'] = pd.to_numeric(df['count'], errors='coerce')
df = df.sort_values(by=['count'], ascending=False)
df = df.sort_values(by=['area', 'cluster', 'cluster_name'])
df = df.reset_index(drop=True)

# Exporting the dataframe into an Excel table:
excel_location = 'sorted_data.xlsx'
writer = pd.ExcelWriter(excel_location, engine='xlsxwriter')
df.to_excel(writer, sheet_name='Sheet1', startrow=1, header=False, index=False)

workbook = writer.book
worksheet = writer.sheets['Sheet1']

# Pinning the headers of the table:
(max_row, max_col) = df.shape
column_settings = [{'header': column} for column in df.columns]
worksheet.add_table(0, 0, max_row, max_col - 1, {'columns': column_settings})

# Making the columns wider for clarity:
worksheet.set_column(0, 0, 12)
worksheet.set_column(2, 2, 15)
worksheet.set_column(3, 3, 40)
worksheet.set_column(6, 6, 22)

writer.save()

# Coloring the text:
file_path = excel_location
wb = load_workbook(file_path)
ws = wb.active

colors = ['17becf', 'bcbd22', '7f7f7f', 'e377c2']
for cell, in ws[f'D2:D{len(df) + 1}']:
    cluster = df['cluster'].iloc[cell.row - 2]
    cell_color = colors[int(cluster)]
    cell.font = Font(color=cell_color)
    color_cell = ws[f'H{cell.row}']
    color_cell.value = df['color'].loc[cell.row - 2] = f'#{cell_color}'

wb.save(file_path)





# PART 2: MAKING GRAPHS

import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
plt.style.use('seaborn-whitegrid')

area_x = []
area_y = []
area_c = []
area_labels = []
area_prev = df['area'].iloc[0]
plot_number = 1

# A function for making separate scatter plots:
def make_plot(plot_num):
    plt.figure(plot_num)
    plt.scatter(x=area_x, y=area_y, c=area_c, s=250, edgecolors='black', linewidths=1)
    for j in range(len(area_labels)):
        plt.annotate(area_labels[j], (area_x[j], area_y[j]), fontsize=7)
    plt.title(area_prev, y=-0.01, fontweight='bold')
    plt.grid(False)
    plt.axis('off')
    color_handles = []
    for j in range(len(colors)):
        color_handles.append(mpatches.Patch(color=f'#{colors[j]}', label=f'Кластер {j}'))
    plt.legend(handles=color_handles, title='Кластеры', frameon=True)
    plt.tight_layout()
    plt.savefig('plot_{plot_num}.png', bbox_inches='tight',dpi=500)

# Making a scatter plot for each area:    
for i in range(len(df)):
    area_now = df['area'].iloc[i]
    if (area_now!=area_prev):
        make_plot(plot_number)
        plot_number += 1
        area_x = []
        area_y = []
        area_c = [] 
        area_labels = []
    area_x.append(df['x'].iloc[i])
    area_y.append(df['y'].iloc[i])
    area_c.append(df['color'].iloc[i])
    marker_label = df['keyword'].iloc[i].replace(' ', '\n')
    area_labels.append(marker_label)
    area_prev = area_now
make_plot(plot_number)
